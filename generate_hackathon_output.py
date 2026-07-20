import csv
import os
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

ZONE2_DIR = "outputs/zone2"
OUTPUT_DIR = "outputs/final_output"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "hackathon_submission.xlsx")

COUNTRIES = ["Singapore", "Malaysia", "Australia"]

TEMPLATE_INDICATORS = OrderedDict([
    ("P6-I1", "General prohibition / restriction"),
    ("P6-I2", "Adequacy standard"),
    ("P6-I3", "Contractual safeguards"),
    ("P6-I4", "Consent exception"),
    ("P6-I5", "Other exceptions"),
    ("P7-I1", "Legal basis for processing"),
    ("P7-I2", "Purpose limitation"),
    ("P7-I3", "Data subject rights"),
    ("P7-I4", "Data breach notification"),
    ("P7-I5", "Enforcement & penalties"),
])

OUR_TO_TEMPLATE = {
    "6.1": "P6-I1",
    "6.2": "P6-I5",
    "6.3": "P6-I5",
    "6.4": "P6-I4",
    "6.5": "P6-I5",
    "7.1": "P7-I1",
    "7.2": "P7-I5",
    "7.3": "P7-I3",
    "7.4": "P7-I1",
    "7.5": "P7-I5",
}

FIELD_ORDER = [
    "Economy",
    "Law Name",
    "Law Number / Ref",
    "Last Amended",
    "Indicator ID",
    "Article / Section",
    "Discovery Tag",
    "Location Reference",
    "Verbatim Snippet",
    "Mapping Rationale",
    "Source URL",
    "Confidence",
    "Notes",
]

PLACEHOLDER_ROWS = {
    ("Singapore", "P6-I2"): {
        "Law Name": "PDPA Section 26 (Transfer Limitation Obligation)",
        "Law Number / Ref": "Act 26 of 2012",
        "Last Amended": "2020",
        "Article / Section": "Section 26",
        "Discovery Tag": "KNOWN",
        "Location Reference": "",
        "Verbatim Snippet": "Indicator P6-I2 (Adequacy standard) is not extracted as a separate provision. The PDPA's Section 26 transfer restriction is enforced via prescribed requirements — where the destination's data protection regime is assessed as part of the organisational obligation to ensure comparable protection.",
        "Mapping Rationale": "Adequacy is implicit in Singapore's conditional transfer model: organisations must verify comparable protection, which functions de facto as adequacy assessment.",
        "Source URL": "https://laws.sg/legislation/personal-data-protection-act-2012",
        "Confidence": "",
        "Notes": "Not extracted as standalone provision. Derived from conditional transfer regime (P6-I4).",
    },
    ("Singapore", "P6-I3"): {
        "Law Name": "PDPA Section 26 (Transfer Limitation Obligation)",
        "Law Number / Ref": "Act 26 of 2012",
        "Last Amended": "2020",
        "Article / Section": "Section 26",
        "Discovery Tag": "KNOWN",
        "Location Reference": "",
        "Verbatim Snippet": "Indicator P6-I3 (Contractual safeguards) is not extracted as a separate provision. Under Singapore's PDPA, organisations may use contractual clauses as a mechanism to ensure comparable protection for cross-border transfers.",
        "Mapping Rationale": "Contractual safeguards are an accepted transfer mechanism under the PDPA transfer limitation obligation. Organisations commonly use BCRs or SCCs to satisfy the comparable protection requirement.",
        "Source URL": "https://laws.sg/legislation/personal-data-protection-act-2012",
        "Confidence": "",
        "Notes": "Not extracted as standalone provision. Derived from conditional transfer regime (P6-I4).",
    },
    ("Singapore", "P6-I5"): {
        "Law Name": "Not extracted (auto-skip indicator)",
        "Law Number / Ref": "",
        "Last Amended": "",
        "Article / Section": "",
        "Discovery Tag": "KNOWN",
        "Location Reference": "",
        "Verbatim Snippet": "Indicator P6-I5 (Other exceptions) covers lawful bases for cross-border transfer beyond consent. The PDPA provides exceptions via prescribed regulations — organisations may transfer where regulations specify permitted mechanisms (SCCs, BCRs, certifications).",
        "Mapping Rationale": "Pipeline limitation: this is a non-regulatory indicator (RDTII 6.5 - binding agreements) scored via external databases. Not extracted through legislation text.",
        "Source URL": "",
        "Confidence": "",
        "Notes": "Auto-skipped in extraction. Scored via external databases in RDTII rubric.",
    },
    ("Malaysia", "P6-I2"): {
        "Law Name": "PDPA 2010 Section 129(2)",
        "Law Number / Ref": "Act 709",
        "Last Amended": "",
        "Article / Section": "Section 129(2)",
        "Discovery Tag": "KNOWN",
        "Location Reference": "",
        "Verbatim Snippet": "For the purposes of subsection (1), the Minister may specify any place outside Malaysia if (a) there is in that place in force any law which is substantially similar to this Act, or that serves the same purposes as this Act; or (b) that place ensures an adequate level of protection in relation to the processing of personal data which is at least equivalent to the level of protection afforded by this Act.",
        "Mapping Rationale": "Section 129(2) explicitly establishes an adequacy standard: transfers are permitted only to jurisdictions deemed by the Minister to have substantially similar laws or adequate protection levels.",
        "Source URL": "https://www.pdp.gov.my/ppdpv1/wp-content/uploads/2024/07/UNDANG-UNDANG-MALAYSIA_AKTA_PERLINDUNGAN_DATA_PERIBADI_2010_709_MALAY_AND-ENG_V2022.pdf",
        "Confidence": "high",
        "Notes": "Subsection referenced within the Section 129 extraction (P6-I4).",
    },
    ("Malaysia", "P6-I3"): {
        "Law Name": "PDPA 2010 Section 129(3)(f)",
        "Law Number / Ref": "Act 709",
        "Last Amended": "",
        "Article / Section": "Section 129(3)(f)",
        "Discovery Tag": "KNOWN",
        "Location Reference": "",
        "Verbatim Snippet": "Notwithstanding subsection (1), a data user may transfer any personal data to a place outside Malaysia if the data user has taken all reasonable precautions and exercised all due diligence to ensure that the personal data will not in that place be processed in any manner which, if that place is Malaysia, would be a contravention of this Act.",
        "Mapping Rationale": "Section 129(3)(f) functions as a contractual safeguard equivalent — a data user may transfer if they exercise due diligence ensuring the recipient handles data consistent with the Act, which is typically achieved through contractual obligations.",
        "Source URL": "https://www.pdp.gov.my/ppdpv1/wp-content/uploads/2024/07/UNDANG-UNDANG-MALAYSIA_AKTA_PERLINDUNGAN_DATA_PERIBADI_2010_709_MALAY_AND-ENG_V2022.pdf",
        "Confidence": "high",
        "Notes": "Subsection referenced within the Section 129 extraction (P6-I4). Due diligence provision serves as contractual safeguard mechanism.",
    },
    ("Australia", "P6-I2"): {
        "Law Name": "Privacy Act 1988 — APP 8 Guidelines",
        "Law Number / Ref": "No. 119, 1988",
        "Last Amended": "2025",
        "Article / Section": "APP 8",
        "Discovery Tag": "KNOWN",
        "Location Reference": "",
        "Verbatim Snippet": "Indicator P6-I2 (Adequacy standard) is implicit in Australia's APP 8 accountability model. Rather than a whitelist, the entity must take reasonable steps to ensure the overseas recipient complies with the APPs — effectively requiring an adequacy assessment per transfer.",
        "Mapping Rationale": "Australia uses an accountability-based model rather than an adequacy determination framework. Entities assess recipient capability on a case-by-case basis.",
        "Source URL": "https://www.oaic.gov.au/privacy/australian-privacy-principles/australian-privacy-principles-guidelines/chapter-8-app-8-cross-border-disclosure-of-personal-information",
        "Confidence": "",
        "Notes": "Not extracted as standalone provision. Derived from APP 8.1 extraction.",
    },
    ("Australia", "P6-I3"): {
        "Law Name": "Privacy Act 1988 — APP 8 Guidelines",
        "Law Number / Ref": "No. 119, 1988",
        "Last Amended": "2025",
        "Article / Section": "APP 8",
        "Discovery Tag": "KNOWN",
        "Location Reference": "",
        "Verbatim Snippet": "Indicator P6-I3 (Contractual safeguards) is implicit in Australia's APP 8.1 requirement to take 'reasonable steps' — entities typically use contractual clauses to ensure the overseas recipient handles information consistently with the APPs.",
        "Mapping Rationale": "Contractual safeguards are the primary mechanism used to satisfy APP 8.1's reasonable steps requirement, though the Act does not prescribe specific model clauses.",
        "Source URL": "https://www.oaic.gov.au/privacy/australian-privacy-principles/australian-privacy-principles-guidelines/chapter-8-app-8-cross-border-disclosure-of-personal-information",
        "Confidence": "",
        "Notes": "Not extracted as standalone provision. Derived from APP 8.1 extraction.",
    },
}

PLACEHOLDER_P7_I2 = {
    "Law Name": "Not extracted (pipeline limitation)",
    "Law Number / Ref": "",
    "Last Amended": "",
    "Article / Section": "",
    "Discovery Tag": "KNOWN",
    "Location Reference": "",
    "Verbatim Snippet": "Indicator P7-I2 (Purpose limitation) evaluates whether data is restricted to the purpose for which it was collected. This indicator was not extracted as a standalone provision in the automated pipeline. It is typically covered under general data protection principles within comprehensive frameworks.",
    "Mapping Rationale": "Pipeline limitation: purpose limitation is not one of the seven extraction indicators in the RDTII scoring rubric used by this pipeline.",
    "Source URL": "",
    "Confidence": "",
    "Notes": "Not extracted. Purpose limitation is evaluated in scoring via the broader data protection framework assessment.",
}

PLACEHOLDER_P7_I4 = {
    "Law Name": "Not extracted (pipeline limitation)",
    "Law Number / Ref": "",
    "Last Amended": "",
    "Article / Section": "",
    "Discovery Tag": "KNOWN",
    "Location Reference": "",
    "Verbatim Snippet": "Indicator P7-I4 (Data breach notification) evaluates mandatory breach notification requirements. This indicator was not extracted as a standalone provision in the automated pipeline run.",
    "Mapping Rationale": "Pipeline limitation: data breach notification is not one of the seven extraction indicators in the RDTII scoring rubric used by this pipeline.",
    "Source URL": "",
    "Confidence": "",
    "Notes": "Not extracted. Breach notification obligations may exist within the broader framework but were not independently extracted.",
}

PLACEHOLDER_P6_I5_COMMON = {
    "Law Name": "Not extracted (auto-skip indicator)",
    "Law Number / Ref": "",
    "Last Amended": "",
    "Article / Section": "",
    "Discovery Tag": "KNOWN",
    "Location Reference": "",
    "Verbatim Snippet": "Indicator P6-I5 (Other exceptions) covers additional lawful bases for cross-border transfer such as vital interests, public interest, and legal proceedings exceptions. Not extracted as standalone provisions.",
    "Mapping Rationale": "Pipeline limitation: other exceptions are assessed within the broader conditional flow regime evaluation or via external databases.",
    "Source URL": "",
    "Confidence": "",
    "Notes": "Auto-skipped or not independently extracted. Refer to P6-I4 conditional regime extraction for available exceptions.",
}

COUNTRY_COLORS = {
    "Singapore": "6366F1",
    "Malaysia": "B45309",
    "Australia": "059669",
}

TITLE_BG = "1E293B"
SUBTITLE_BG = "1D4ED8"
REQUIRED_BG = "059669"
OPTIONAL_BG = "94A3B8"
INSTRUCTION_BG = "F1F5F9"
INSTRUCTION_TEXT = "64748B"
BODY_TEXT = "1E293B"
BORDER_COLOR = "E2E8F0"
DATA_ALT_BG = "F8FAFC"
WHITE = "FFFFFF"
PLACEHOLDER_FILL = "FFFBEB"


def truncate(text, max_chars=300):
    if not text:
        return ""
    text = str(text).strip()
    if len(text) <= max_chars:
        return text
    return text[:max_chars-3] + "..."


def col_fill(rgb):
    return PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")


def font_f(name="Calibri", size=10, bold=False, color=BODY_TEXT):
    return Font(name=name, size=size, bold=bold, color=color)


def align(horizontal="left", vertical="center", wrap=True):
    return Alignment(horizontal=horizontal, vertical=vertical, wrap_text=wrap)


THIN = Side(style="thin", color=BORDER_COLOR)
MEDIUM = Side(style="medium", color=BORDER_COLOR)
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_MEDIUM_TOP = Border(
    left=THIN, right=THIN, top=MEDIUM, bottom=THIN,
)
BORDER_NONE = Border()


def read_zone2_csv(filepath):
    with open(filepath, encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def build_real_rows(all_rows):
    seen = set()
    rows_by_country = {c: {} for c in COUNTRIES}
    for row in all_rows:
        economy = (row.get("Economy") or "").strip()
        if economy not in rows_by_country:
            continue
        indicator_id = (row.get("Indicator_ID") or "").strip()
        mapped_id = OUR_TO_TEMPLATE.get(indicator_id)
        if not mapped_id:
            continue

        law_name = (row.get("Act_and_or_practice") or "").strip()
        article_section = (row.get("Article_Section") or "").strip()

        dedup_key = (economy, law_name, article_section, mapped_id)
        if dedup_key in seen:
            continue
        seen.add(dedup_key)

        out = {
            "Economy": economy,
            "Law Name": law_name,
            "Law Number / Ref": (row.get("Law_Number_Ref") or "").strip(),
            "Last Amended": (row.get("Last_Amended") or "").strip(),
            "Indicator ID": mapped_id,
            "Article / Section": article_section,
            "Discovery Tag": (row.get("Discovery_Tag") or "").strip(),
            "Location Reference": (row.get("Location_Reference") or "").strip(),
            "Verbatim Snippet": (row.get("Verbatim_Snippet") or "").strip(),
            "Mapping Rationale": truncate((row.get("Mapping_Rationale") or "").strip(), 300),
            "Source URL": (row.get("References") or "").strip(),
            "Confidence": (row.get("Confidence") or "").strip(),
            "Notes": _build_notes(row, indicator_id, mapped_id),
            "_is_placeholder": False,
        }
        if mapped_id not in rows_by_country[economy]:
            rows_by_country[economy][mapped_id] = []
        rows_by_country[economy][mapped_id].append(out)

    return rows_by_country


def _build_notes(row, orig_id, mapped_id):
    parts = []
    timeframe = (row.get("Timeframe") or "").strip()
    if timeframe:
        parts.append(timeframe)
    coverage = (row.get("Coverage") or "").strip()
    if coverage and coverage != "Cross-cutting":
        parts.append(f"Cvg: {coverage}")
    return " | ".join(parts)


def ensure_all_indicators(rows_by_country):
    for country in COUNTRIES:
        for tid in TEMPLATE_INDICATORS:
            pillar = tid.split("-")[0]
            number = int(tid.split("-")[1][1])
            if tid not in rows_by_country[country]:
                placeholder = None
                if tid == "P7-I2":
                    ph = dict(PLACEHOLDER_P7_I2)
                    ph["Economy"] = country
                    ph["Indicator ID"] = tid
                    placeholder = ph
                elif tid == "P7-I4":
                    ph = dict(PLACEHOLDER_P7_I4)
                    ph["Economy"] = country
                    ph["Indicator ID"] = tid
                    placeholder = ph
                elif (country, tid) in PLACEHOLDER_ROWS:
                    ph = dict(PLACEHOLDER_ROWS[(country, tid)])
                    ph["Economy"] = country
                    ph["Indicator ID"] = tid
                    placeholder = ph
                else:
                    ph = dict(PLACEHOLDER_P6_I5_COMMON)
                    ph["Economy"] = country
                    ph["Indicator ID"] = tid
                    placeholder = ph
                if placeholder:
                    placeholder["_is_placeholder"] = True
                    rows_by_country[country][tid] = [placeholder]


def build_output_sorted(rows_by_country):
    out = []
    for country in COUNTRIES:
        country_data = rows_by_country[country]
        for tid in TEMPLATE_INDICATORS:
            if tid not in country_data:
                continue
            for row in country_data[tid]:
                out.append(row)
    return out


def write_header(wb):
    ws = wb.active
    ws.title = "Output Data"

    widths = {
        "A": 18, "B": 34, "C": 20, "D": 14, "E": 14, "F": 18,
        "G": 14, "H": 22, "I": 58, "J": 46, "K": 46, "L": 11, "M": 34,
    }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    r1 = ws.cell(row=1, column=1, value="Team PillarAI - UN ESCAP Global Hackathon - Final Output")
    r1.font = Font(name="Calibri", size=16, bold=True, color=WHITE)
    r1.fill = col_fill(TITLE_BG)
    r1.alignment = align("center", "center")
    ws.merge_cells("A1:M1")
    ws.row_dimensions[1].height = 32

    r2 = ws.cell(row=2, column=1, value="Round 1 Submission  |  Indicators 6 & 7 (Cross-border Data Flows + Domestic Data Protection)")
    r2.font = Font(name="Calibri", size=10, bold=False, color=WHITE)
    r2.fill = col_fill(SUBTITLE_BG)
    r2.alignment = align("center", "center")
    ws.merge_cells("A2:M2")
    ws.row_dimensions[2].height = 20

    for i, col_name in enumerate(FIELD_ORDER, start=1):
        cell = ws.cell(row=3, column=i, value=col_name)
        cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        cell.fill = col_fill(REQUIRED_BG)
        cell.alignment = align("center", "center")
        cell.border = BORDER_THIN

    instr = [
        "Economy name (use official UN name)",
        "Full official statute name and year",
        "Official law/act number",
        "Year of most recent amendment",
        'RDTII indicator code (P6-I1..P7-I5)',
        "Exact article and paragraph",
        'NEW or KNOWN',
        "PDF page or HTML anchor",
        "Exact verbatim text",
        "Max 300 chars. Maps to indicator because...",
        "Direct URL to official government portal",
        "Certainty score (0.00-1.00)",
        "Flag unusual cases, OCR issues",
    ]
    for i, text in enumerate(instr, start=1):
        cell = ws.cell(row=4, column=i, value=text)
        cell.font = Font(name="Calibri", size=8, bold=False, color=INSTRUCTION_TEXT)
        cell.fill = col_fill(INSTRUCTION_BG)
        cell.alignment = align("left", "center")
        cell.border = BORDER_THIN
    ws.row_dimensions[4].height = 36

    ws.row_dimensions[3].height = 22

    return ws


def write_country_header(ws, row_num, country):
    bg = COUNTRY_COLORS.get(country, TITLE_BG)
    for i in range(1, 14):
        cell = ws.cell(row=row_num, column=i)
        cell.fill = col_fill(bg)
        cell.border = Border(
            left=THIN, right=THIN,
            top=Side(style="medium", color=bg),
            bottom=Side(style="medium", color=bg),
        )
    c1 = ws.cell(row=row_num, column=1, value=country.upper())
    c1.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    c1.alignment = align("left", "center")
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=13)
    ws.row_dimensions[row_num].height = 30
    return row_num + 1


def write_data_rows(ws, start_row, data_rows):
    row_num = start_row
    current_country = None
    current_indicator = None

    for d in data_rows:
        country = d["Economy"]
        indicator = d["Indicator ID"]
        is_placeholder = d.get("_is_placeholder", False)
        is_last_in_group = False

        if country != current_country:
            current_country = country
            current_indicator = None
            row_num = write_country_header(ws, row_num, country)

        if indicator != current_indicator:
            current_indicator = indicator

        b_fill = None
        if is_placeholder:
            b_fill = col_fill(PLACEHOLDER_FILL)
        elif (row_num % 2) == 0:
            b_fill = col_fill(DATA_ALT_BG)

        for col_idx, col_name in enumerate(FIELD_ORDER, start=1):
            val = d.get(col_name, "")
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = Font(name="Calibri", size=10,
                             bold=False,
                             color=INSTRUCTION_TEXT if is_placeholder else BODY_TEXT)
            cell.alignment = align("left", "center")
            cell.border = BORDER_THIN
            if b_fill:
                cell.fill = b_fill

            if col_name == "Indicator ID":
                cell.font = Font(name="Calibri", size=10, bold=True,
                                 color=BODY_TEXT)
            if col_name == "Discovery Tag":
                tag = str(val).strip()
                if tag == "NEW":
                    cell.font = Font(name="Calibri", size=10, bold=True,
                                     color="059669")
                elif tag == "KNOWN":
                    cell.font = Font(name="Calibri", size=10, bold=False,
                                     color="6366F1")

        ws.row_dimensions[row_num].height = 65
        row_num += 1

    return row_num


def write_indicator_reference(wb):
    ws = wb.create_sheet("Indicator Reference", 1)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 65

    r1 = ws.cell(row=1, column=1, value="RDTII Indicator Reference — Pillars 6 & 7")
    r1.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    r1.fill = col_fill(TITLE_BG)
    r1.alignment = align("left", "center")
    for c in range(2, 4):
        ws.cell(row=1, column=c).fill = col_fill(TITLE_BG)
    ws.merge_cells("A1:C1")

    ref_data = [
        ("Pillar 6: Cross-border Data Flows", None, None),
        ("P6-I1", "General prohibition / restriction",
         "Does the law restrict cross-border transfer as a default?"),
        ("P6-I2", "Adequacy standard",
         "Can data be transferred to countries deemed to have adequate protection?"),
        ("P6-I3", "Contractual safeguards",
         "Are SCCs or BCRs accepted as transfer mechanisms?"),
        ("P6-I4", "Consent exception",
         "Can transfer proceed with individual consent?"),
        ("P6-I5", "Other exceptions",
         "What other lawful bases permit cross-border transfer?"),
        ("Pillar 7: Domestic Data Protection", None, None),
        ("P7-I1", "Legal basis for processing",
         "Does the law require a lawful basis for processing?"),
        ("P7-I2", "Purpose limitation",
         "Is data restricted to the purpose collected?"),
        ("P7-I3", "Data subject rights",
         "Do individuals have access, correction, deletion rights?"),
        ("P7-I4", "Data breach notification",
         "Is there mandatory breach notification?"),
        ("P7-I5", "Enforcement & penalties",
         "Is there a supervisory authority and penalty regime?"),
    ]

    for i, (a, b, c) in enumerate(ref_data, start=2):
        cell_a = ws.cell(row=i, column=1, value=a)
        cell_b = ws.cell(row=i, column=2, value=b)
        cell_c = ws.cell(row=i, column=3, value=c)

        is_header = b is None and c is None
        if is_header:
            hdr = "2563EB" if "6:" in (a or "") else "059669"
            cell_a.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
            cell_a.fill = col_fill(hdr)
            cell_a.alignment = align("left", "center")
            for c2 in range(2, 4):
                ws.cell(row=i, column=c2).fill = col_fill(hdr)
                ws.cell(row=i, column=c2).border = BORDER_THIN
        else:
            for cell in [cell_a, cell_b, cell_c]:
                cell.font = Font(name="Calibri", size=10,
                                 bold=(cell.column == 2),
                                 color=BODY_TEXT)
                cell.alignment = align("left", "center")
                cell.border = BORDER_THIN

        cell_a.border = BORDER_THIN
        cell_b.border = BORDER_THIN
        cell_c.border = BORDER_THIN

    return ws


def print_summary(data_rows, all_raw_count):
    by_country = {}
    by_indicator = {}
    placeholder_count = 0
    real_count = 0
    for d in data_rows:
        c = d["Economy"]
        ind = d["Indicator ID"]
        pip = d.get("_is_placeholder", False)
        if pip:
            placeholder_count += 1
        else:
            real_count += 1
        by_country[c] = by_country.get(c, 0) + 1
        by_indicator[ind] = by_indicator.get(ind, 0) + 1

    print(f"\n  Total outputs: {len(data_rows)} rows")
    print(f"  Real extractions: {real_count} | Placeholder rows: {placeholder_count}")
    print(f"  Countries:")
    for c in COUNTRIES:
        print(f"    {c}: {by_country.get(c, 0)} rows")
    print(f"  Indicators:")
    for tid in TEMPLATE_INDICATORS:
        cnt = by_indicator.get(tid, 0)
        label = TEMPLATE_INDICATORS[tid][:50]
        print(f"    {tid}: {cnt} rows  ({label})")


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    all_rows = []
    for fname in sorted(os.listdir(ZONE2_DIR)):
        if not fname.endswith(".csv"):
            continue
        rows = read_zone2_csv(os.path.join(ZONE2_DIR, fname))
        all_rows.extend(rows)
        print(f"  Read {len(rows)} rows from {fname}")

    rows_by_country = build_real_rows(all_rows)
    ensure_all_indicators(rows_by_country)
    data_rows = build_output_sorted(rows_by_country)

    wb = Workbook()
    ws = write_header(wb)
    write_data_rows(ws, 5, data_rows)
    write_indicator_reference(wb)

    wb.save(OUTPUT_FILE)
    print_summary(data_rows, len(all_rows))
    print(f"\n  Wrote: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
